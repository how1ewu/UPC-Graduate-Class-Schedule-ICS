import datetime
import openpyxl
import re
import icalendar
import uuid
import copy

workbook = openpyxl.load_workbook('./test.xlsx')
sheet = workbook.active

events = [] 

# 定义获取事件文本的函数  
def get_event(row, col, cell, merge_count):
    event = str(row-1) + ' ' + str(col-1) + ' ' 
    event += str(merge_count) + ' '
    event += str(cell.value)
    return event

# 获取所有合并单元格范围
merged_cells = sheet.merged_cells.ranges  

# 遍历所有单元格
merged_cells = sheet.merged_cells

for row in range(2, sheet.max_row+1):
  for col in range(2, sheet.max_column+1):  
    
    cell = sheet.cell(row=row, column=col)  
    if cell.value is not None:
      is_merged = False
      for merged_cell in merged_cells:
        if row >= merged_cell.min_row and row <= merged_cell.max_row and col >= merged_cell.min_col and col <= merged_cell.max_col:
          mrow = merged_cell.max_row - merged_cell.min_row + 1
          mcol = merged_cell.max_col - merged_cell.min_col + 1
          is_merged = True
          break
        
      if is_merged:
        events.append(get_event(row, col, cell, mrow))
      else:
        events.append(get_event(row, col, cell, 1))
result = []

for item in events:
    parts = re.split(r'\s+', item, maxsplit=3)
    time = parts[0]  
    day = parts[1]
    mount = parts[2]
    course_info = parts[3]

    course_name = re.search(r'^.+?(?=\{)', course_info).group()  
    weeks_info = re.search(r'\{(.+?)\}', course_info).group(1)

    teacher, local = None, None
    _teacher = re.search(r'\[教师:(.+?)\,', weeks_info)
    _local = re.search(r'\,地点:(.+?)\]', weeks_info)
    if _teacher:
        teacher = _teacher.group(1)
        local = _local.group(1)
    if ';' in weeks_info:
        for weekly_course in weeks_info.split(';'):
          if '、' in weekly_course and '-' in weekly_course:
            weeks = re.split('-', weekly_course)
            range_weeks = weeks[0] + '-' + weeks[1]
            single_weeks = weeks[0].split('、')
            for single in single_weeks:
                if re.search(r'\[教师:(.+?)\]', single):
                  week, teacher = re.search(r'(\d+)周\[教师:(.+)\]',  single).groups()
                  result.append([time, day, mount, course_name, week, teacher, local])
            start_week = re.split('、', weeks[0])[-1]
            end_week = re.search(r'(\d+)', weeks[1]).group(1)
            teacher = None
            if re.search(r'\[教师:(.+?)\]', weeks[1]):
              teacher = re.search(r'\[教师:(.+?)\]', weeks[1]).group(1)
            for week in range(int(start_week), int(end_week)+1):
                result.append([time, day, mount, course_name, str(week), teacher, local])
          elif '、' in weekly_course:
            for weekly_info in weekly_course.split('、'):
                week, teacher = re.search(r'(\d+)周\[教师:(.+)\]', weekly_info).groups()
                result.append([time, day, mount, course_name, week, teacher, local])  

          else:
            weeks = weekly_course.split('-')
            start_week = re.search(r'(\d+)', weeks[0]).group(1)
            end_week = re.search(r'(\d+)', weeks[1]).group(1)
            teacher = None
            if re.search(r'\[教师:(.+?),', weekly_course):
                teacher = re.search(r'\[教师:(.+?),', weekly_course).group(1)

            for week in range(int(start_week), int(end_week)+1):
                result.append([time, day, mount, course_name, str(week), teacher, local])

    else:
      if '、' in weeks_info and '-' in weeks_info:
          weeks = re.split('-', weeks_info)
          range_weeks = weeks[0] + '-' + weeks[1]
          single_weeks = weeks[0].split('、')
          for single in single_weeks:
              if re.search(r'\[教师:(.+?)\]', single):
                week, teacher = re.search(r'(\d+)周\[教师:(.+)\]',  single).groups()
                result.append([time, day, mount, course_name, week, teacher, local])
          start_week = re.split('、', weeks[0])[-1]
          end_week = re.search(r'(\d+)', weeks[1]).group(1)
          teacher = None
          if re.search(r'\[教师:(.+?)\]', weeks[1]):
            teacher = re.search(r'\[教师:(.+?)\]', weeks[1]).group(1)
          for week in range(int(start_week), int(end_week)+1):
              result.append([time, day, mount, course_name, str(week), teacher, local])
      elif '、' in weeks_info:
          for weekly_info in weeks_info.split('、'):
              week, teacher = re.search(r'(\d+)周\[教师:(.+)\]', weekly_info).groups()
              result.append([time, day, mount, course_name, week, teacher, local])  

      else:
          weeks = weeks_info.split('-')
          start_week = re.search(r'(\d+)', weeks[0]).group(1)
          end_week = re.search(r'(\d+)', weeks[1]).group(1)
          teacher = None
          if re.search(r'\[教师:(.+?),', weeks_info):
              teacher = re.search(r'\[教师:(.+?),', weeks_info).group(1)

          for week in range(int(start_week), int(end_week)+1):
              result.append([time, day, mount, course_name, str(week), teacher, local])


def get_date(start_year, start_month, start_day, start_weekday, week_num, week_day):
    """
    start_year:第一周的起始年
    start_month: 第一周的起始月
    start_day: 第一周的起始日
    start_weekday: 第一周的起始星期(0-6,0为星期一)
    week_num: 已过第几周
    week_day: 第几天(0-6,0为星期一)
    """ 
    first_day = datetime.date(start_year, start_month, start_day)
    start_delta = (start_weekday - first_day.weekday()) % 7
    target_delta = (week_num - 1) * 7 + (week_day - start_weekday) % 7
    return first_day + datetime.timedelta(days=start_delta + target_delta)

# 设置日期计算的起始点和一周的第一天 
start_date = datetime.date(2023, 9, 3)
first_weekday = 6 

# 计算日期
# target_date = start_date + datetime.timedelta(days=(week-1)*7 + weekday - first_weekday)
cal = icalendar.Calendar(tzid='Asia/Shanghai')
time_slots = {
    1: (datetime.time(8,0), datetime.time(8,45)),
    2: (datetime.time(8,50), datetime.time(9,35)),
    3: (datetime.time(9,55), datetime.time(10,40)),
    4: (datetime.time(10,45), datetime.time(11,30)),
    5: (datetime.time(11,35), datetime.time(12,20)),
    6: (datetime.time(14,0), datetime.time(14,45)),
    7: (datetime.time(14,50), datetime.time(15,35)),
    8: (datetime.time(15,55), datetime.time(16,40)),
    9: (datetime.time(16,45), datetime.time(17,30)),
    10: (datetime.time(19,0), datetime.time(19,45)),
    11: (datetime.time(19,50), datetime.time(20,35)),
    12: (datetime.time(20,40), datetime.time(21,25))
}
weekdays = {
   1:6,
   2:0,
   3:1,
   4:2,
   5:3,
   6:4,
   7:5  
}
for event in result:
    start_slot = int(event[0])
    end_slot = int(event[0])+int(event[2])-1
    duration = int(event[1])
    
    start_time = time_slots[start_slot][0]
    end_time = time_slots[end_slot][1]
    
    weekday = int(event[4])
    day = int(weekdays[int(event[1])])
    target_date = get_date(2023, 9, 3, 6, weekday, day) 
    #target_date = start_date + datetime.timedelta(days=(weekday-1)*7 + day - first_weekday)
    
    event_name = event[3]
    instructor = event[5]
    location = event[6]
    uid=uuid.uuid4()
    evt = icalendar.Event()
    evt.add('summary', event_name)
    evt.add('dtstart', datetime.datetime.combine(target_date, start_time))
    evt.add('dtend', datetime.datetime.combine(target_date, end_time))
    evt.add('location', location)
    evt.add('description', instructor)
    evt.add('uid', uid)
    
    cal.add_component(evt)

# print(cal.to_ical().decode('utf-8'))
merged_events = {}
merged_cal = icalendar.Calendar()
merged_event_uids = set()

for event in cal.subcomponents:
    evt = copy.deepcopy(event)   
    day = evt['DTSTART'].dt.date()
            
    if day not in merged_events:
        merged_events[day] = evt
    else:
        existing_event = merged_events[day]
        if (evt['DTSTART'].dt - existing_event['DTEND'].dt).total_seconds() < 21*60:
            merged_events[day] = icalendar.Event()
            merged_events[day].add('SUMMARY', existing_event['SUMMARY'])
            merged_events[day].add('DESCRIPTION', existing_event['DESCRIPTION'])
            merged_events[day].add('LOCATION', existing_event['LOCATION'])
            merged_events[day].add('DTSTART', min(existing_event['DTSTART'].dt, evt['DTSTART'].dt))
            merged_events[day].add('DTEND', max(existing_event['DTEND'].dt, evt['DTEND'].dt))
            merged_events[day].add('UID', str(uuid.uuid4()))
            merged_event_uids.add(evt['UID'])
            merged_event_uids.add(existing_event['UID'])
            
for event in merged_events.values():
    merged_cal.add_component(event)
        
for event in cal.subcomponents:
  if event['UID'] not in merged_event_uids:
    merged_cal.add_component(event)
# 保存文件
with open('./calendar.ics', 'wb') as f:
  f.write(merged_cal.to_ical())

print("日程表已保存到 calendar.ics")